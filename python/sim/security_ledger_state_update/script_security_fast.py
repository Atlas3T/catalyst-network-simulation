import numpy
import random
import math
import shutil
import argparse
from itertools import chain
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import os.path
import time
from random import randint
from time import sleep
import multiprocessing as mp
import glob 
import pandas as pd
import zipfile


def if_test_pass(ind_p, process_id, results_test, runs_full, runs_test, spec):
    results_full = numpy.array([calculate_lists_rate(**spec) for _ in range(runs_full)])
    
    results = numpy.concatenate((results_test, results_full))
    runs = runs_full + runs_test
    outputs = get_result_output(spec['num_of_producers'],
                                spec['prop_correct_producers'],
                                runs=runs, results=results)
    pid = str(os.getpid())
    path_name = "excel/Result_simulation_security_ledger_update" + pid + ".xlsx"
    write_results_to_excel_file(spec, runs=runs, output=outputs, process_id=process_id,
                                path_name=path_name) #If this fails we need to retry 
        
    prob_it = numpy.count_nonzero(results[:, 1] > ind_p / 2) / runs
    print(f"P = {ind_p}, "
            f"prod = {spec['prop_correct_producers']}, "
            f"update = {spec['prop_collected_update']}, "
            f"vote = {spec['prop_collected_candidate']}, "
            f"final vote = {spec['prop_collected_vote']} --> {prob_it}")
    return prob_it


def run_experiment_hist(spec, step_producer, end_producer, step_sce, end_sce, step_prop, end_prop, runs_test, runs_full):
    process_id = os.getpid()
    start_time = time.time()
    start_producer = spec['num_of_producers']
    start_sce = spec['prop_correct_producers']
    start_prop = spec['prop_collected_update']
    for ind_p in range(start_producer, end_producer, step_producer):
        spec['num_of_producers'] = ind_p
        spec['prop_correct_producers'] = start_sce
        while spec['prop_correct_producers'] < end_sce:

            spec['prop_collected_update'] = start_prop
            spec['prop_collected_candidate'] = start_prop
            spec['prop_collected_vote'] = start_prop
            prob_it = 0
            while spec['prop_collected_update'] < end_prop:
                results_test = numpy.array([calculate_lists_rate(**spec) for _ in range(runs_test)])
                test_pass = numpy.count_nonzero(results_test[:, 1] > ind_p / 2) / runs_test
                
                print("P = ",ind_p, ", a=", spec['prop_correct_producers'], ", b=",
                        spec['prop_collected_update'], ", c=", spec['prop_collected_candidate'],
                        ", d=", spec['prop_collected_vote'], " -->", test_pass, "-->", process_id)
                if test_pass >= 0.6:
                 
                    prob_it = if_test_pass(ind_p, process_id, results_test, runs_full, runs_test, spec)
                   
                spec['prop_collected_update'] *= 100
                spec['prop_collected_update'] += step_prop*100
                spec['prop_collected_update'] /= 100
                #spec['prop_collected_update'] = int(spec['prop_collected_update']*10000)/10000
                spec['prop_collected_candidate'] = spec['prop_collected_update']
                spec['prop_collected_vote'] = spec['prop_collected_update']

                if prob_it > 0.999:
                    spec['prop_collected_update'] = end_prop
                    spec['prop_collected_candidate'] = spec['prop_collected_update']
                    spec['prop_collected_vote'] = spec['prop_collected_update']
        
            #spec['prop_correct_producers'] += step_sce
            spec['prop_correct_producers'] *= 100
            spec['prop_correct_producers'] += step_sce*100
            spec['prop_correct_producers'] /= 100
            spec['prop_correct_producers'] = int(spec['prop_correct_producers']*10000)/10000
            
                     
def run_experiment_grad(spec, step_producer, end_producer, step_prop, end_prop, runs_test, runs_full):

    start_producer = spec['num_of_producers']
    start_prop = spec['prop_correct_producers']
    for ind_p in range(start_producer, end_producer, step_producer):
        list_pass_params = numpy.array([[1, 1, 1, 1]])
        spec['num_of_producers'] = ind_p
        spec['prop_correct_producers'] = start_prop
        while spec['prop_correct_producers'] < end_prop:
            spec['prop_collected_update'] = start_prop
            while spec['prop_collected_update'] < end_prop:
                spec['prop_collected_candidate'] = start_prop
                while spec['prop_collected_candidate'] < end_prop:
                    spec['prop_collected_vote'] = start_prop
                    while spec['prop_collected_vote'] < end_prop:
                        list_temp_a = list_pass_params[
                            (list_pass_params[:, 0] < spec['prop_correct_producers']) &
                            (list_pass_params[:, 1] <= spec['prop_collected_update']) &
                            (list_pass_params[:, 2] <= spec['prop_collected_candidate']) &
                            (list_pass_params[:, 3] <= spec['prop_collected_vote'])]

                        list_temp_b = list_pass_params[
                            (list_pass_params[:, 0] == spec['prop_correct_producers']) &
                            (list_pass_params[:, 1] < spec['prop_collected_update']) &
                            (list_pass_params[:, 2] <= spec['prop_collected_candidate']) &
                            (list_pass_params[:, 3] <= spec['prop_collected_vote'])]

                        list_temp_c = list_pass_params[
                            (list_pass_params[:, 0] == spec['prop_correct_producers']) &
                            (list_pass_params[:, 1] == spec['prop_collected_update']) &
                            (list_pass_params[:, 2] < spec['prop_collected_candidate']) &
                            (list_pass_params[:, 3] <= spec['prop_collected_vote'])]

                        list_temp_d = list_pass_params[
                            (list_pass_params[:, 0] == spec['prop_correct_producers']) &
                            (list_pass_params[:, 1] == spec['prop_collected_update']) &
                            (list_pass_params[:, 2] == spec['prop_collected_candidate']) &
                            (list_pass_params[:, 3] < spec['prop_collected_vote'])]

                        if len(list_temp_a) == 0 and len(list_temp_b) == 0 and len(list_temp_c) == 0 \
                                and len(list_temp_d) == 0:
                            results_test = numpy.array([calculate_lists_rate(**spec) for _ in range(runs_test)])
                            test_pass = numpy.count_nonzero(results_test[:, 1] > ind_p / 2) / runs_test

                            print("a=", spec['prop_correct_producers'], ", b=",
                                  spec['prop_collected_update'], ", c=", spec['prop_collected_candidate'],
                                  ", d=", spec['prop_collected_vote'], " -->", test_pass)
                            if test_pass > 0.2:
                                results_full = numpy.array([calculate_lists_rate(**spec) for _ in range(runs_full)])

                                results = numpy.concatenate((results_test, results_full))
                                runs = runs_full + runs_test
                                outputs = get_result_output(spec['num_of_producers'],
                                                            spec['prop_correct_producers'],
                                                            runs=runs, results=results)
                                write_results_to_excel_file(spec, runs=runs, output=outputs,
                                                            path_name="Result_simulation_security_bla.xlsx")
                                full_pass = numpy.count_nonzero(results[:, 1] > ind_p / 2) / runs
                                if numpy.count_nonzero(results[:, 1] > ind_p / 2) == runs:
                                    set_params = [[spec['prop_correct_producers'],
                                                  spec['prop_collected_update'],
                                                  spec['prop_collected_candidate'],
                                                  spec['prop_collected_vote']]]
                                    list_pass_params = numpy.concatenate((list_pass_params, set_params))
                                    print(list_pass_params)
                                print(f"P = {ind_p}, "
                                      f"prod = {spec['prop_correct_producers']}, "
                                      f"update = {spec['prop_collected_update']}, "
                                      f"vote = {spec['prop_collected_candidate']}, "
                                      f"final vote = {spec['prop_collected_vote']} --> {full_pass}")

                        spec['prop_collected_vote'] += step_prop
                        spec['prop_collected_vote'] = int(spec['prop_collected_vote']*10000)/10000
                    spec['prop_collected_candidate'] += step_prop
                    spec['prop_collected_candidate'] = int(spec['prop_collected_candidate']*10000)/10000
                spec['prop_collected_update'] += step_prop
                spec['prop_collected_update'] = int(spec['prop_collected_update']*10000)/10000
            spec['prop_correct_producers'] += step_prop
            spec['prop_correct_producers'] = int(spec['prop_correct_producers']*10000)/10000


# Functions for output results
def get_result_output(num_of_producers, prop_correct_producers, runs, results):
    avg_prod = numpy.mean(results[:, 0]) / num_of_producers
    avg_vote = numpy.mean(results[:, 1]) / num_of_producers
    num_list100_prod = (1 - numpy.count_nonzero(results[:, 0] - prop_correct_producers * num_of_producers) / runs)
    num_list100_vote = (1 - numpy.count_nonzero(results[:, 1] - num_of_producers) / runs)
    num_pass = numpy.count_nonzero(results[:, 1] > num_of_producers / 2) / runs

    return {
        'avg_prod': avg_prod,
        'avg_vote': avg_vote,
        'num_list100_prod': num_list100_prod,
        'num_list100_vote': num_list100_vote,
        'num_pass': num_pass
    }


def print_result_output(num_of_producers, prop_correct_producers, prop_collected_update, prop_collected_candidate,
                               prop_collected_vote, runs, output):

    print(f'Parameters: P = {num_of_producers} ({runs} runs)')
    print(f'PropCorrectProducer = {prop_correct_producers*100}%')
    print(f'PropCollectedUpdate = {prop_collected_update*100}%')
    print(f'PropCollectedCandidate = {prop_collected_candidate*100}%')
    print(f'PropCollectedVote = {prop_collected_vote*100}%')
    print('############################')
    print('Averages:')
    print(f'{output["avg_prod"]*100:10.3f} % producers issue correct Ln(prod)')
    print(f'{output["avg_vote"]*100:10.3f} % producers issue correct Ln(vote)')
    print('Successful runs:')
    print(f'{output["num_list100_prod"]*100:10.3f} % runs with no missing data for Ln(prod)')
    print(f'{output["num_list100_vote"]*100:10.3f} % runs with no missing data for Ln(vote)')
    print('Summary:')
    print(f'{output["num_pass"]*100:10.3f} % successful runs ( > 50% producers broadcast same update)')
    print('############################')


def initiate_worksheet(workbook, sheet_title):
    sheet_result = workbook.worksheets[workbook.index(workbook[sheet_title])]
    ind_col = 1
    if sheet_result['A1'].value is None:
        name_cols = ['Total Producers', '% Correct Producers', '% Collected Updates', '% Collected Votes',
                     '% Collected Final Votes', 'runs', 'Avg. % producers issue correct Ln(prod)',
                     'Avg. % producers issue correct Ln(vote)', '% runs with no missing data for Ln(prod)',
                     '% runs with no missing data for Ln(vote)',
                     '% successful runs w/ > 50% producers broadcast same update','ID of process that carried out work (Ignore for single runs)']
        for name_col in name_cols:
            sheet_result.cell(row=1, column=ind_col).value = name_col
            sheet_result.column_dimensions[get_column_letter(ind_col)].width = 20
            sheet_result.cell(row=1, column=ind_col).alignment = Alignment(wrap_text=True)
            ind_col += 1
    return sheet_result





def write_results_to_excel_file(spec, runs, output, process_id, path_name):

    
    wb = Workbook()
    if os.path.isfile(path_name):
        wb = load_workbook(path_name)

    sheet_result_title = "P_" + str(spec['num_of_producers'])
    #+  + "_"
    if sheet_result_title not in wb.sheetnames:
        wb.create_sheet(sheet_result_title)

    sheet_result = initiate_worksheet(wb, sheet_result_title)
    ind_col = 1
    ind_row = 1
    while sheet_result.cell(row=ind_row, column=ind_col).value is not None:
        ind_row += 1
    # parameters
    for col_exc in spec:
        sheet_result.cell(row=ind_row, column=ind_col).value = spec[col_exc]
        ind_col += 1
    # runs
    sheet_result.cell(row=ind_row, column=ind_col).value = runs
    ind_col += 1
    # output
    for col_out in output:
        sheet_result.cell(row=ind_row, column=ind_col).value = output[col_out]
        ind_col += 1
    #processID
    sheet_result.cell(row=ind_row, column=ind_col).value = process_id
    ind_col += 1
   

    wb.save(path_name)


# Functions for calculate_lists_rate
def get_list_producer_ids(num_producers, num_sample):
    list_ids = numpy.zeros((num_producers, num_sample), int)
    for i in range(num_producers):
        list_ids[i, 0] = i
        list_ids[i, 1:num_sample] = random.sample(list(chain(range(0, i), range(i + 1, num_producers))), num_sample - 1)
    return list_ids


def get_list_producers_who_found_majority(num_producer, id_correct_producers, collected_data_ids, num_collected_data):
    majority_found = numpy.zeros(num_producer, int)
    for i in range(num_producer):
        k_maj = numpy.count_nonzero(id_correct_producers[collected_data_ids[i, :]] == 1)
        ratio_maj = k_maj/num_collected_data
        #print(num_collected_data,", ",ratio_maj)
        ratio_threshold = 0.5 + 4.22 * math.sqrt(ratio_maj * (1 - ratio_maj) / num_collected_data)
        majority_found[i] = 1 if ratio_maj > ratio_threshold else 0
    return majority_found


def get_partial_list_ids_found(num_producer, id_correct_producers, collected_data_ids, num_collected_data):

    is_majority_found = get_list_producers_who_found_majority(num_producer, id_correct_producers, collected_data_ids,
                                                           num_collected_data)

    partial_list = numpy.zeros((num_producer, num_collected_data), int)
    for i in range(num_producer):
        partial_list[i, :] = -1
        if is_majority_found[i] == 1:
            for j in range(num_collected_data):
                if id_correct_producers[collected_data_ids[i, j]] == 1:
                    partial_list[i, j] = collected_data_ids[i, j]
                else:
                    partial_list[i, j] = -1
    return partial_list


def get_merged_list_ids(num_producer, sampled_lists, merging_threshold):
    merged_list = []
    for i in range(num_producer):
        unique, counts = numpy.unique(sampled_lists[i], return_counts=True)
        correct_producers = dict(zip(unique, counts))
        l_temp = []
        for producer, count in correct_producers.items():
            if count > merging_threshold and producer != -1:
                l_temp.append(producer)
        merged_list.append(l_temp)
    return merged_list


def get_ids_with_full_lists(num_producer, correct_list, lists_to_compare):
    count_list_complete = []
    for i in range(num_producer):
        if lists_to_compare[i] == correct_list:
            count_list_complete.append(i)
    return count_list_complete


def calculate_lists_rate(num_of_producers, prop_correct_producers, prop_collected_update, prop_collected_candidate,
                         prop_collected_vote):
    """
    This function checks how many producers manage to compile the correct lists Ln(prod) and Ln(vote)
    :param num_of_producers: Number of producers (P)
    :param prop_correct_producers: Fraction of producers who correctly build the dominant ledger state update (C_n/P)
    :param prop_collected_update: Fraction of collected ledger state update per producer (C_j/P)
    :param prop_collected_candidate: Fraction of collected candidate per producer (V_j/P)
    :param prop_collected_vote: Fraction of collected vote per producer (U_j/C_n)
    :return: result (int, int) where the first is the fraction of producers among K_n having correctly built Ln(prod)
    and the second is the fraction of producers among P having correctly built Ln(vote)
    """
    result = numpy.zeros(2, int)

    # Range to experiment [0.55,0.99]
    num_of_correct_updates = math.floor(prop_correct_producers * num_of_producers)

    # Producers flag: 1 if associated to correct update, 0 otherwise
    correct_update_flags = numpy.zeros(num_of_producers, int)
    correct_update_flags[:num_of_correct_updates] = 1

    # True list of correct producers (K_n)
    correct_update_ids = [i for i in range(num_of_correct_updates)]

    # K_j updates collected per producer. Range to experiment [0.75,0.99], use to determine k_min
    num_of_collected_updates = math.floor(prop_collected_update*num_of_producers)

    # List of producer ids associated to updates collected by producer
    collected_update_ids = get_list_producer_ids(num_of_producers, num_of_collected_updates)

    # List of id of producers who found a dominant update
    majority_found = get_list_producers_who_found_majority(num_of_producers, correct_update_flags, collected_update_ids,
                                                           num_of_collected_updates)

    # List of producer ids associated to correct updates collected by producer, -1 otherwise
    lj_prod = get_partial_list_ids_found(num_of_producers, correct_update_flags, collected_update_ids,
                                         num_of_collected_updates)

    # True list of correct voters
    correct_vote_ids = []
    for i in range(num_of_producers):
        if majority_found[i] == 1:
            correct_vote_ids.append(i)
    correct_vote_ids.sort()

    if len(correct_vote_ids) == 0:
        print("No producer found a majority of same update: abort!")
        return result

    # V_j votes collected per producer. Range to experiment [0.75,0.99]. Use to determine V_min
    num_of_collected_candidate = math.floor(prop_collected_candidate * num_of_producers)

    # List of producer ids associated to votes collected by producer
    collected_vote_ids = get_list_producer_ids(num_of_producers, num_of_collected_candidate)
    collected_vote_ids = collected_vote_ids[0:num_of_correct_updates, :]

    # Set of lj_prod collected by each producer
    lj_prod_sampled = numpy.zeros((num_of_correct_updates, num_of_collected_candidate, num_of_collected_updates), int)
    for i in range(num_of_correct_updates):
        lj_prod_sampled[i, :, :] = [lj_prod[collected_vote_ids[i, j], :] for j in range(num_of_collected_candidate)]

    # List of producers associated to correct update merged by each producers (compare with correct_update_ids)
    compare_correct_update_ids = get_merged_list_ids(num_of_correct_updates, lj_prod_sampled, num_of_producers/2)

    producer_ids_full_list_prod = get_ids_with_full_lists(num_of_correct_updates, correct_update_ids,
                                                          compare_correct_update_ids)

    # List of producer ids associated to correct vote collected by producer, -1 otherwise
    lj_vote = numpy.zeros((num_of_correct_updates, num_of_collected_candidate), int)
    for i in range(num_of_correct_updates):
        majority_i = get_list_producers_who_found_majority(num_of_collected_candidate, correct_update_flags,
                                                           lj_prod_sampled[i, :, :], num_of_collected_updates)

        lj_vote[i, :] = [collected_vote_ids[i, k] if majority_i[k] == 1 and k in producer_ids_full_list_prod else -1
                         for k in range(num_of_collected_candidate)]

    # U_j <=  num_of_correct_updates final votes collected per producer. Range to experiment [0.75,0.99]
    num_of_collected_votes = math.floor(prop_collected_vote * num_of_correct_updates)

    # Set of lj_vote collected by each producer
    lj_vote_sampled = numpy.zeros((num_of_producers, min(num_of_collected_votes, num_of_collected_candidate),
                                   num_of_collected_candidate), int)
    threshold_vote_list = len(correct_update_ids) / 2

    for i in range(num_of_producers):
        lj_vote_sampled[i, :, :] = lj_vote[numpy.random.choice(lj_vote.shape[0], min(num_of_collected_votes,
                                                                                     num_of_collected_candidate),
                                                               replace=False), :]
        if numpy.count_nonzero(lj_vote_sampled[i, :] != -1) < threshold_vote_list:
            lj_vote_sampled[i, :, :] = -1

    compare_correct_vote_producer_ids = get_merged_list_ids(num_of_producers, lj_vote_sampled, threshold_vote_list)

    producer_ids_full_list_vote = get_ids_with_full_lists(num_of_producers, correct_vote_ids,
                                                          compare_correct_vote_producer_ids)

    result = [len(producer_ids_full_list_prod), len(producer_ids_full_list_vote)]

    return result


def parse_args():
    parser = argparse.ArgumentParser(description='Process some integers.')
    parser.add_argument('--p', type=int, default=200, help='Number of producers')
    parser.add_argument('--runs', type=int, default=10, help='Number of runs')
    parser.add_argument('--producer', type=float, default=0.8, help='Proportion of correct producers')
    parser.add_argument('--update', type=float, default=0.8, help='Proportion of collected updates per producer')
    parser.add_argument('--candidate', type=float, default=0.8, help='Proportion of collected candidate per producer')
    parser.add_argument('--vote', type=float, default=0.8, help='Proportion of collected votes')
    return parser.parse_args()


def setup_spec():
    spec = {
        'num_of_producers': 100,
        'prop_correct_producers': 0.9,
        'prop_collected_update': 0.85,
        'prop_collected_candidate': 0.85,
        'prop_collected_vote': 0.85
    }
    step_producer = 100
    end_producer = 201
    step_sce = 0.1
    end_sce = 0.91
    step_prop = 0.01
    end_prop = 0.96
    
    spec_test = spec.copy()
    run_test = 5
    run_full = 5
    list_pass_test = []

    return (end_producer, end_prop, end_sce, 
    list_pass_test, run_full, run_test, 
    spec, spec_test, step_producer, 
    step_prop, step_sce)


def combine_excel_files(end_producer, step_producer, spec):
    glob.glob("excel/*.xlsx")
    all_data = pd.DataFrame()
    start_producer = spec['num_of_producers']
    
    for f in glob.glob("excel/*.xlsx"):
        for ind_p in range(start_producer, end_producer, step_producer):
            ind_p_str = str(ind_p)
            sheetID = str(ind_p_str)
            df = pd.read_excel(f, "P_" + sheetID)
            all_data = all_data.append(df,ignore_index=True) 
        
    
    print(all_data['Total Producers'](index=False))
    all_data.to_excel("output.xlsx", sheet_name="P_"+all_data['Total Producers'].to_string(index=False))

     
def move_old_excel():
    timestr = get_time()
    os.rename('excel','old_excel/excel_'+timestr)
    os.mkdir('excel')
    

def run_multiprocessing(spec_test, step_producer, end_producer, step_sce, end_sce, step_prop, end_prop, run_test, run_full):
    for num in range(2):
        mp.Process(target=run_experiment_hist, args=(spec_test, step_producer, end_producer, step_sce, end_sce, step_prop, end_prop, run_test, run_full)).start()


def get_time():
    return time.strftime("%Y%m%d-%H%M%S")


if __name__ == '__main__':
    
    step_producer = 0
    end_producer = 0
    step_sce = 0
    end_sce = 0
    step_prop = 0
    end_prop = 0
    
    spec_test = {}
    run_test = 0
    run_full = 0
    list_pass_test = 0
    spec = {}
    
    
    level_test = 4
    
    if level_test == 0:
        print(" No test selected")

    if level_test == 1:
        args = parse_args()
        res = numpy.array([calculate_lists_rate(args.p, args.producer, args.update, 
                                                args.candidate, args.vote)
                           for _ in range(args.runs)])
        output = get_result_output(args.p, args.producer, 
                                   runs=args.runs, results=res)

        print_result_output(args.p, args.producer, 
                                   args.update, args.candidate, 
                                   args.vote, runs=args.runs,
                                   output=output)
    if level_test == 2:
        end_producer, end_prop, end_sce, 
        list_pass_test, run_full, run_test, 
        spec, spec_test, step_producer, 
        step_prop, step_sce = setup_spec()

        run_experiment_grad(spec_test, step_producer, end_producer, step_prop, end_prop, run_test, run_full)

    if level_test == 3: 
        (end_producer, end_prop, end_sce, 
        list_pass_test, run_full, run_test, 
        spec, spec_test, step_producer, 
        step_prop, step_sce) = setup_spec()

        move_old_excel()

        run_experiment_hist(spec_test, step_producer, end_producer, step_sce, end_sce, step_prop, end_prop, run_test, run_full)
        timestr = get_time()
        os.rename('excel','level_3_single_runs/excel_'+timestr) # Saves individual file to its own folder under 'level_3_single_runs'
        os.mkdir('excel')

    if level_test == 4: #This test is the same as above but allows multiprocessing
        (end_producer, end_prop, end_sce, 
        list_pass_test, run_full, run_test, 
        spec, spec_test, step_producer, 
        step_prop, step_sce) = setup_spec()
        
        #Removes any old excel files from previous runs of the script 
        move_old_files = mp.Process(target=move_old_excel,args=()) 
        move_old_files.start()
        move_old_files.join()
        
        #Runs the multiprocessing of the run_experiment_hist
        run_multiprocessing_runs = mp.Process(target=run_multiprocessing, args=(spec_test, step_producer, end_producer, step_sce, end_sce, step_prop, end_prop, run_test, run_full))
        run_multiprocessing_runs.start()
        run_multiprocessing_runs.join()

        combine_excel = mp.Process(combine_excel_files(end_producer, step_producer, spec))    
        print("Combining excel files") 
        combine_excel.start()
        combine_excel.join()