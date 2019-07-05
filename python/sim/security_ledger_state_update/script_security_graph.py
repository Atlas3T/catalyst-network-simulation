from anytree import Node, RenderTree
from anytree.exporter import DotExporter
import graphviz
from openpyxl import load_workbook
import numpy as np

def check_excel_sheets_readable(path_name="Result_simulation_security_hist-test.xlsx",sheet_ranges=""): 
    try: 
        wb = load_workbook(filename)
        try:
            source = wb.get_sheet_by_name(sheet_ranges)
            return True
        except KeyError as e:
            print("Unable to open sheet {}".format(sheet_ranges))
            return False
    except IOError as e:
        print("Unable to open workbook {}".format(path_name))
        return False


def worksheet_to_array(path_name="Result_simulation_security_hist-test.xlsx"):
    wb = load_workbook(filename)
    a = ["P_100","P_200"]
    data = []
    for i in a:
        source = wb.get_sheet_by_name(i)
        firstRow = 2
        firstCol = 0
        nCols = 100
        nRows = 100
        allCells = np.array([[cell.value for cell in row] for row in source.iter_rows()])
        data.append(allCells[(firstRow-1):(firstRow-1+nRows),(firstCol-1):(firstCol-1+nCols)])
    return(data)


def ext_P_values():
    data = data.worksheet_to_array

if __name__ == '__main__':
    filename = 'Result_simulation_security_hist-test.xlsx'
    data = []
    data = worksheet_to_array(data)
    print (data)
