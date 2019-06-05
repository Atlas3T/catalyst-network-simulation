from anytree import Node, RenderTree
from anytree.exporter import DotExporter
import graphviz
from openpyxl import load_workbook
import numpy as np


def check_excel_sheets_readable(path_name="Result_simulation_security_ledger_update.xlsx",sheet_ranges=""): 
    try: 
        wb = load_workbook(filename)
        try:
            source = wb.get_sheet_by_name(sheet_ranges)
            return True
        except KeyError as e:
            print("Unable to open sheet {}".format(sheet_ranges))
            return False
    except IOError as e:
        print("Unable to open workbook {}".format(path_name))
        return False


def worksheet_to_array(path_name="Result_simulation_security_ledger_update.xlsx",sheet_ranges=""):
    wb = load_workbook(filename)
    source = wb.get_sheet_by_name(sheet_ranges)
    firstRow = 2
    firstCol = 2
    nCols = 100
    nRows = 100
    allCells = np.array([[cell.value for cell in row] for row in source.iter_rows()])
    data = allCells[(firstRow-1):(firstRow-1+nRows),(firstCol-1):(firstCol-1+nCols)]
    return(data)


if __name__ == '__main__':


    #This function checks that both the excel file and the excel sheets can be opened#  
    filename = 'test.xlsx'
    node_p = "P_100"
    if check_excel_sheets_readable(filename, node_p) is not True:
         print("abort")
         exit()

    #Here we extract the cells from the array
    Full_Array = worksheet_to_array(filename,node_p)
    # print(Full_Array)

    node_p_tree = Node(node_p)

    #C_n, C_min, V_min, U_min
    unique_cn = np.unique(Full_Array[:,0], return_counts=False) #Produces a list of unique I'Ds for first collumn
    print("Cn -- ", unique_cn)
    unique_cn_id = 0 
    unique_cn_val = unique_cn[unique_cn_id]
    unique_cn_it = unique_cn_val #From those unique values it selects the first value from that list 
    std_cn = "Cn: " + unique_cn_it
    node_cn_it = Node("{}".format(std_cn), parent=node_p_tree) #Creates the top level i.e. p=100
    # cretae the first child of the first parent

    
    unique_cmin = np.unique(Full_Array[Full_Array[:,0] == unique_cn[unique_cn_id],1], return_counts=False) #Produces a list of unique I'Ds for first collumn
    print("Cmin -- ", unique_cmin)
    unique_cmin_id = 0 
    unique_cmin_val = unique_cmin[unique_cmin_id]
    unique_cmin_it = unique_cmin_val #From those unique values it selects the first value from that list 
    std_cmin = "Cmin: " + unique_cmin_it
    node_cmin_it = Node("{}".format(std_cmin), parent=node_cn_it) #Creates the top level i.e. p=100

    #unique_vmin = np.unique(Full_Array([(Full_Array[:,0] == unique_cn[unique_cn_id],1) and (Full_Array[:,1] == unique_cmin[unique_cmin_id]),2], return_counts=False))  #Produces a list of unique I'Ds for first collumn
    unique_vmin = np.unique(Full_Array[Full_Array[:,1] == unique_cmin[unique_cmin_id],2], return_counts=False)
    print("Vmin -- ", unique_vmin)
    unique_vmin_id = 0 
    #unique_vmin_val = unique_vmin[unique_vmin_id]
    unique_vmin_val = unique_vmin[unique_vmin_id]
    unique_vmin_it = unique_vmin_val #From those unique values it selects the first value from that list 
    std_vmin = "Vmin: " + unique_vmin_it
    node_vmin_it = Node("{}".format(std_vmin), parent=node_cmin_it) #Creates the top level i.e. p=100


    #create the parent node 1
    for rowa in Full_Array: #for each row in the full array 

        if rowa[0] == unique_cn_val:
            if rowa[1] == unique_cmin_val: #if first row == to the parent ID
                print("same")
            
            elif rowa[2] == unique_vmin_val:
                print("same")
            else:
                unique_vmin_id += 1
                unique_vmin_val = unique_vmin[unique_vmin_id]
                unique_vmin_it = unique_vmin_val #From those unique values it selects the first value from that list 
                std_vmin = "Vmin: " + unique_vmin_it
                node_vmin_it = Node("{}".format(std_vmin), parent=node_cmin_it)
                
        elif rowa[0] == unique_cn_val: #if first row == to the parent ID
            #creatre a child 
            if rowa[1] == unique_cmin_val:
                print("same")
            else:
                unique_cmin_id += 1
                unique_cmin_val = unique_cmin[unique_cmin_id]
                unique_cmin_it = unique_cmin_val #From those unique values it selects the first value from that list 
                std_cmin = "Cmin: " + unique_cmin_it
                node_cmin_it = Node("{}".format(std_cmin), parent=node_cn_it) #Creates the top level i.e. p=100
        else:
            unique_cn_id += 1
            print(unique_cn_id)
            unique_cn_val = unique_cn[unique_cn_id]
            unique_cn_it = unique_cn_val
            std_cn = "Cn: " + unique_cn_it
            node_cn_it  = Node("{}".format(std_cn), parent=node_p_tree)

            unique_cmin = np.unique(Full_Array[Full_Array[:,0] == unique_cn[unique_cn_id],1], return_counts=False) #Produces a list of unique I'Ds for first collumn
            print("Cmin -- ", unique_cmin)
            unique_cmin_id = 0 
            unique_cmin_val = unique_cmin[unique_cmin_id]
            unique_cmin_it = unique_cmin_val #From those unique values it selects the first value from that list 
            std_cmin = "Cmin: " + unique_cmin_it
            node_cmin_it = Node("{}".format(std_cmin), parent=node_cn_it) #Creates the top level i.e. p=100

            
            unique_vmin = np.unique(Full_Array[Full_Array[:,1] == unique_cmin[unique_cmin_id],2], return_counts=False)   #Produces a list of unique I'Ds for first collumn
            print("Vmin -- ", unique_vmin)
            unique_vmin_id = 0 
            unique_vmin_val = unique_vmin[unique_vmin_id]
            unique_vmin_it = unique_vmin_val #From those unique values it selects the first value from that list 
            std_vmin = "Vmin: " + unique_vmin_it
            node_vmin_it = Node("{}".format(std_vmin), parent=node_cmin_it) #Creates the top level i.e. p=100

       
    for pre, fill, node in RenderTree(node_p_tree):
        print("%s%s" % (pre, node.name)) 

    #    while Full_Array[i,0] == unique_a_it:
    #        print(i, " ",Full_Array[i,0])
    #    unique_a_it = Full_Array[i,0]

    #DotExporter(udo).to_picture("udo.png")

    #main that execute

    #def can you read from excel --> True


    #def read from excel --> array (a,b,c,d)


#def build the tree (one per P)


